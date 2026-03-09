# ##report_generators.py
from openpyxl.styles import Font
from typing import Optional
import openpyxl
from datetime import datetime, date
from telegram import Update
from telegram.error import Forbidden
from telegram.ext import ContextTypes
import os
import logging

from database import db
from config import CONFIG
from models import User, Order
from sqlalchemy import text

from report_utils import ensure_reports_dir
from settings import SETTINGS_CONFIG

logger = logging.getLogger(__name__)

async def export_orders_for_provider(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
):
    """
    Формирует текстовый отчёт для поставщиков.
    Объединяет данные по локациям "Офис" и "ПЦ 2".
    Добавляет строку с незарегистрированными сотрудниками, если такие есть.
    """
    try:
        # Если даты не заданы — используем сегодняшнюю
        if not start_date or not end_date:
            today = datetime.now(CONFIG.timezone).date()
            start_date = end_date = today
        else:
            start_date = start_date if isinstance(start_date, date) else start_date.date()
            end_date = end_date if isinstance(end_date, date) else end_date.date()

        with db.get_session() as session:
            # Получаем данные по локациям
            location_data_result = session.execute(text('''
                SELECT COALESCE(u.location, 'Не указано'), SUM(o.quantity)
                FROM orders o
                JOIN users u ON o.user_id = u.id
                WHERE o.target_date BETWEEN :start_date AND :end_date
                  AND o.is_cancelled = FALSE
                GROUP BY COALESCE(u.location, 'Не указано')
            '''), {'start_date': start_date.isoformat(), 'end_date': end_date.isoformat()})

            location_data = dict(location_data_result.fetchall())

        # Объединяем "Офис" и "ПЦ 2"
        office_portions = location_data.get("Офис", 0) + location_data.get("ПЦ 2", 0)
        pc1_portions = location_data.get("ПЦ 1", 0)
        warehouse_portions = location_data.get("Склад", 0)
        unregistered_portions = location_data.get("Не указано", 0)

        total_portions = office_portions + pc1_portions + warehouse_portions + unregistered_portions

        # Формируем текстовое сообщение
        period_text = (
            f"{start_date.strftime('%d.%m.%Y')}"
            if start_date == end_date
            else f"{start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
        )

        # Основные строки отчета
        message_lines = [
            f"📋 *Заказы на* | {period_text}",
            f"━━━━━━━━━━━━━━━━━━",
            f"📌 Всего: *{total_portions}* порций\n",
            f"• 🏢 Офис: *{office_portions}*",
            f"• 🏭 ПЦ 1: *{pc1_portions}*",
            f"• 📦 Склад: *{warehouse_portions}*"
        ]

        # Добавляем строку с незарегистрированными только если они есть
        if unregistered_portions > 0:
            message_lines.append(f"• ❓ Незарегистрированные: *{unregistered_portions}*")

        # Собираем финальное сообщение
        message = "\n".join(message_lines)

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=message,
            parse_mode="Markdown"
        )

    except Forbidden:
        raise
    except Exception as e:
        logger.error(f"Ошибка при создании отчета для поставщика: {e}", exc_info=True)
        try:
            await update.message.reply_text("❌ Ошибка при формировании отчёта.")
        except Forbidden:
            pass
        raise
    
async def export_accounting_report(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
):
    try:
        # 1. Настройка форматирования чисел
        import locale
        try:
            locale.setlocale(locale.LC_NUMERIC, 'ru_RU.UTF-8')
        except:
            pass

        # 2. Словарь месяцев
        month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
        }

        reports_dir = ensure_reports_dir('accounting')
        now = datetime.now(CONFIG.timezone)
        
        # Обработка дат
        if not start_date or not end_date:
            start_date = end_date = now.date()
        else:
            start_date = start_date if isinstance(start_date, date) else start_date.date()
            end_date = end_date if isinstance(end_date, date) else end_date.date()

        if start_date > end_date:
            start_date, end_date = end_date, start_date

        # Определяем месяц и год по периоду отчета
        report_month = start_date.month
        report_year = start_date.year
        month_year = f"{month_names[report_month]} {report_year}"

        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Удержания за обеды"
        
        # Заголовки
        ws.append(["Список сотрудников на удержание обедов из ежемесячной премии"])
        ws.append([f"за {month_year} г."])
        ws.append([])
        ws.append(["", "удержание стоимости 1 обеда составляет", "150,00 руб. (без НДФЛ)"])
        ws.append(["", "", "172,41 руб. (с НДФЛ 13%)"])
        ws.append([])
        
        # Заголовки таблицы
        headers = [
            "Подразделение",
            "ФИО",
            "Кол-во обедов",
            "Должность",
            "Территория",
            "Дата приема",
            "Сумма удержания без НДФЛ",
            "Сумма удержания с НДФЛ"
        ]
        ws.append(headers)
        ws.auto_filter.ref = f"A{ws.max_row}:H{ws.max_row}"

        # 🔥 ИСПРАВЛЕННЫЙ ЗАПРОС - используем employment_date вместо created_at
        with db.get_session() as session:
            query = text('''
                SELECT 
                    COALESCE(u.department, 'Не указано') as department,
                    u.full_name,
                    SUM(o.quantity) as portions,
                    COALESCE(u.position, 'Не указана') as position,
                    COALESCE(u.city, 'Не указана') as city,
                    u.employment_date as hire_date  -- 🔥 Берем дату приема из employment_date
                FROM orders o
                JOIN users u ON o.user_id = u.id
                WHERE o.target_date BETWEEN :start_date AND :end_date
                  AND o.is_cancelled = FALSE
                GROUP BY u.id, u.department, u.full_name, u.position, u.city, u.employment_date
                ORDER BY u.department, u.full_name
            ''')
            
            rows = session.execute(query, {
                'start_date': start_date,
                'end_date': end_date
            }).fetchall()

        # Инициализация переменных
        total_portions = 0
        total_without_ndfl = 0
        total_with_ndfl = 0

        if not rows:
            ws.append(["Нет данных за выбранный период", "", "", "", "", "", "", ""])
        else:
            for row in rows:
                department, full_name, portions, position, city, hire_date = row
                
                # 🔥 ОБРАБОТКА ДАТЫ ПРИЕМА - преобразуем в читаемый формат
                hire_date_str = "Не указана"
                if hire_date:
                    if isinstance(hire_date, date):
                        hire_date_str = hire_date.strftime("%d.%m.%Y")
                    elif isinstance(hire_date, datetime):
                        hire_date_str = hire_date.date().strftime("%d.%m.%Y")
                    elif isinstance(hire_date, str):
                        try:
                            # Пробуем разные форматы дат
                            if '.' in hire_date:
                                # Формат DD.MM.YYYY
                                hire_date_str = datetime.strptime(hire_date, "%d.%m.%Y").strftime("%d.%m.%Y")
                            elif '-' in hire_date:
                                # Формат YYYY-MM-DD
                                hire_date_str = datetime.strptime(hire_date, "%Y-%m-%d").strftime("%d.%m.%Y")
                            else:
                                hire_date_str = "Не указана"
                        except:
                            hire_date_str = "Не указана"
                
                # Расчет сумм
                amount_without_ndfl = portions * 150
                amount_with_ndfl = round(amount_without_ndfl / 0.87, 2)
                
                ws.append([
                    department,  # подразделение
                    full_name,   # ФИО
                    portions,    # количество обедов
                    position,    # должность
                    city,        # территория (город)
                    hire_date_str,  # 🔥 дата приема из employment_date
                    amount_without_ndfl,  # числовое значение
                    amount_with_ndfl      # числовое значение
                ])
                
                total_portions += portions
                total_without_ndfl += amount_without_ndfl
                total_with_ndfl += amount_with_ndfl

            # Итоговая строка
            ws.append([
                "ВСЕГО",
                "",
                total_portions,
                "",
                "",
                "",
                total_without_ndfl,  # числовое значение
                total_with_ndfl      # числовое значение
            ])

        # Форматирование
        bold_font = Font(bold=True)
        money_format = '# ##0.00'
        
        # Применение стилей
        for row in ws.iter_rows(min_row=1, max_row=6):
            for cell in row:
                cell.font = bold_font
                
        for cell in ws[7]:  # заголовки таблицы
            cell.font = bold_font
            
        for cell in ws[ws.max_row]:  # итоговая строка
            cell.font = bold_font
        
        # Формат денежных значений
        for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
            for cell in row[6:8]:  # колонки с суммами
                if cell.value and isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
                    cell.value = float(cell.value)
        
        # Автоподбор ширины столбцов
        column_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                length = len(str(cell.value)) * 1.2
                if cell.column_letter not in column_widths or length > column_widths[cell.column_letter]:
                    column_widths[cell.column_letter] = length
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = min(width, 50)
        
        # Сохранение файла
        file_name = f"salary_deductions_{report_year}{report_month:02d}.xlsx"
        file_path = os.path.join(reports_dir, file_name)
        wb.save(file_path)
        
        # Функция форматирования валюты для вывода в сообщении
        def format_currency(amount):
            return f"{float(amount):,.2f}".replace(",", " ").replace(".", ",")

        # Отправка файла
        caption = (
            f"📋 Отчет для удержаний из зарплаты\n"
            f"📅 Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n"
            f"🍽 Всего обедов: {total_portions}\n"
            f"💰 Сумма удержания: {format_currency(total_with_ndfl)} руб. (с НДФЛ)"
        )

        with open(file_path, 'rb') as file:
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=file,
                caption=caption,
                filename=file_name
            )

        return file_path

    except Forbidden:
        raise
    except Exception as e:
        logger.error(f"Ошибка формирования отчета: {e}", exc_info=True)
        try:
            await update.message.reply_text("❌ Произошла ошибка при создании отчета.")
        except Forbidden:
            pass
        raise
    
async def export_monthly_report(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    is_daily: bool = False
):
    """Генерация административного отчёта с возможностью указания дат"""
    try:
        if update.effective_user.id not in CONFIG.admin_ids:
            await update.message.reply_text("❌ У вас нет прав для выполнения этой команды.")
            return

        now = datetime.now(CONFIG.timezone)
        
        # Если даты не переданы - используем текущий месяц
        if not start_date or not end_date:
            month_start = now.replace(day=1).date()
            end_date = now.date()
            start_date = month_start
        else:
            # Проверяем, что start_date <= end_date
            if start_date > end_date:
                start_date, end_date = end_date, start_date

        reports_dir = ensure_reports_dir('admin')
        
        wb = openpyxl.Workbook()
        
        # Удаляем лист по умолчанию
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Создаем лист "Все заказы"
        ws_all = wb.create_sheet("Все заказы", 0)
        all_headers = ["Дата обеда", "Номер заказа", "Сотрудник", "Локация", "Подпись", "Кол-во обедов", "Источник заказа"]
        ws_all.append(all_headers)
        ws_all.auto_filter.ref = "A1:G1"
        
        # 🔥 ИСПРАВЛЕННЫЙ ЗАПРОС с SQLAlchemy
        with db.get_session() as session:
            if is_daily:
                query = text("""
                    SELECT 
                        o.target_date,
                        u.full_name,
                        COALESCE(u.location, 'Не указано') as location,
                        o.quantity,
                        o.is_from_bitrix,
                        o.created_at,
                        o.bitrix_order_id
                    FROM orders o
                    JOIN users u ON o.user_id = u.id
                    WHERE o.target_date = :target_date
                    AND o.is_cancelled = FALSE
                    ORDER BY 
                        o.target_date,
                        CASE 
                            WHEN o.bitrix_order_id IS NULL THEN CAST(o.created_at AS TEXT)
                            ELSE o.bitrix_order_id 
                        END,
                        u.full_name
                """)
                result = session.execute(query, {'target_date': start_date})
            else:
                query = text("""
                    SELECT 
                        o.target_date,
                        u.full_name,
                        COALESCE(u.location, 'Не указано') as location,
                        o.quantity,
                        o.is_from_bitrix,
                        o.created_at,
                        o.bitrix_order_id
                    FROM orders o
                    JOIN users u ON o.user_id = u.id
                    WHERE o.target_date BETWEEN :start_date AND :end_date
                    AND o.is_cancelled = FALSE
                    ORDER BY 
                        o.target_date,
                        CASE 
                            WHEN o.bitrix_order_id IS NULL THEN CAST(o.created_at AS TEXT)
                            ELSE o.bitrix_order_id 
                        END,
                        u.full_name
                """)
                result = session.execute(query, {'start_date': start_date, 'end_date': end_date})
            
            all_orders = result.fetchall()

        # 🔥 ПРОВЕРКА: если нет заказов
        if not all_orders:
            period_desc = start_date.strftime("%d.%m.%Y") if is_daily else f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            await update.message.reply_text(f"📊 На период {period_desc} заказов нет")
            return

        # Группируем заказы по дате для общего листа
        orders_by_date = {}
        for row in all_orders:
            date_key = row[0]  # target_date
            if date_key not in orders_by_date:
                orders_by_date[date_key] = []
            orders_by_date[date_key].append(row)
        
        # Заполняем лист "Все заказы" с номером заказа
        for date_key in sorted(orders_by_date.keys()):
            for row in orders_by_date[date_key]:
                target_date = row[0].strftime("%d.%m.%Y") if isinstance(row[0], date) else datetime.strptime(row[0], "%Y-%m-%d").strftime("%d.%m.%Y")
                source = "Битрикс" if row[4] else "Бот"
                order_number = row[6] if row[6] is not None else ""
                ws_all.append([target_date, order_number, row[1], row[2], "", row[3], source])

        # 🔥 ОСТАЛЬНАЯ ЛОГИКА ОСТАЕТСЯ ПРЕЖНЕЙ (создание листов по локациям, итоги и т.д.)
        # Создаем листы для каждой локации (объединяем "Офис" и "ПЦ 2")
        for location in CONFIG.locations:
            if location == "ПЦ 2":
                continue
                
            ws = wb.create_sheet(location)
            headers = ["Дата обеда", "Номер заказа", "Сотрудник", "Территориальный признак", "Подпись", "Кол-во обедов", "Источник заказа"]
            ws.append(headers)
            ws.auto_filter.ref = "A1:G1"
            
            # Для листа "Офис" включаем также заказы из "ПЦ 2"
            if location == "Офис":
                location_orders = [row for row in all_orders if row[2] in ["Офис", "ПЦ 2"]]
            else:
                location_orders = [row for row in all_orders if row[2] == location]
            
            # Группируем по дате для текущей локации
            loc_orders_by_date = {}
            for row in location_orders:
                date_key = row[0]
                if date_key not in loc_orders_by_date:
                    loc_orders_by_date[date_key] = []
                loc_orders_by_date[date_key].append(row)
            
            # Заполняем лист локации с номером заказа
            for date_key in sorted(loc_orders_by_date.keys()):
                for row in loc_orders_by_date[date_key]:
                    target_date = row[0].strftime("%d.%m.%Y") if isinstance(row[0], date) else datetime.strptime(row[0], "%Y-%m-%d").strftime("%d.%m.%Y")
                    source = "Битрикс" if row[4] else "Бот"
                    order_number = row[6] if row[6] is not None else ""
                    ws.append([target_date, order_number, row[1], row[2], "", row[3], source])
        
        # Лист "Итоги"
        ws_summary = wb.create_sheet("Итоги")
        summary_headers = ["Локация", "Порции"]
        ws_summary.append(summary_headers)
        ws_summary.auto_filter.ref = "A1:B1"
        
        # Подсчет итогов (объединяем "Офис" и "ПЦ 2")
        location_totals = {}
        for row in all_orders:
            location = row[2]
            if location == "ПЦ 2":
                location = "Офис"
            quantity = row[3]
            if location not in location_totals:
                location_totals[location] = 0
            location_totals[location] += quantity
        
        total = 0
        for location, portions in sorted(location_totals.items(), key=lambda x: x[1], reverse=True):
            ws_summary.append([location, portions])
            total += portions
        
        ws_summary.append(["ВСЕГО", total])
        
        # Форматирование
        bold_font = Font(bold=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = bold_font
            
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = max_length + 2

        # Сохраняем файл
        timestamp = now.strftime("%Y%m%d_%H%M%S")
        file_name = f"admin_report_{timestamp}.xlsx"
        file_path = os.path.join(reports_dir, file_name)
        wb.save(file_path)
        
        # Формируем сообщение
        if is_daily:
            caption = f"📅 Админ отчет за {start_date.strftime('%d.%m.%Y')}\n🍽 Всего порций: {total}"
        else:
            caption = f"📅 Админ отчет за период {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n🍽 Всего порций: {total}"
        
        with open(file_path, 'rb') as file:
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=file,
                caption=caption,
                filename=file_name
            )

    except Forbidden:
        raise
    except Exception as e:
        logger.error(f"Ошибка формирования админ отчёта: {e}", exc_info=True)
        try:
            await update.message.reply_text("❌ Ошибка формирования отчёта")
        except Forbidden:
            pass
        
async def export_daily_admin_report(
    update: Update, 
    context: ContextTypes.DEFAULT_TYPE, 
    target_date: Optional[date] = None
):
    """Формирует дневной административный отчет"""
    logger.info(f"Создание дневного админ отчета, дата: {target_date}")
    if not target_date:
        target_date = datetime.now(CONFIG.timezone).date()
    return await export_monthly_report(
        update, 
        context, 
        target_date, 
        target_date,
        is_daily=True  # Передаём флаг, что это дневной отчёт
    )
        
    
async def export_daily_orders_for_provider(update: Update, context: ContextTypes.DEFAULT_TYPE, target_date: Optional[date] = None):
    """Формирует дневной отчет для поставщиков"""
    logger.info(f"Создание дневного отчета для поставщика, дата: {target_date}")
    if not target_date:
        target_date = datetime.now(CONFIG.timezone).date()
    return await export_orders_for_provider(update, context, target_date, target_date)