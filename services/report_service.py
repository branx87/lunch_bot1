"""
Report generation business logic: SQL queries + Excel file creation.
Messenger-agnostic — used by both Telegram and Max bots.
"""
import os
import logging
from datetime import datetime, date
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import text

from config import CONFIG
from database import db
from report_utils import ensure_reports_dir

logger = logging.getLogger(__name__)

MONTH_NAMES = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}


def generate_provider_report_text(start_date, end_date, session):
    """
    Generate provider text report (location-based order summary).
    Returns (report_text, total_portions).
    """
    if not start_date or not end_date:
        today = datetime.now(CONFIG.timezone).date()
        start_date = end_date = today
    else:
        start_date = start_date if isinstance(start_date, date) else start_date.date()
        end_date = end_date if isinstance(end_date, date) else end_date.date()

    result = session.execute(text('''
        SELECT COALESCE(u.location, 'Не указано'), SUM(o.quantity)
        FROM orders o
        JOIN users u ON o.user_id = u.id
        WHERE o.target_date BETWEEN :start_date AND :end_date
          AND o.is_cancelled = FALSE
        GROUP BY COALESCE(u.location, 'Не указано')
    '''), {'start_date': start_date.isoformat(), 'end_date': end_date.isoformat()})

    location_data = dict(result.fetchall())

    office_portions = location_data.get("Офис", 0) + location_data.get("ПЦ 2", 0)
    pc1_portions = location_data.get("ПЦ 1", 0)
    warehouse_portions = location_data.get("Склад", 0)
    unregistered_portions = location_data.get("Не указано", 0)
    total_portions = office_portions + pc1_portions + warehouse_portions + unregistered_portions

    period_text = (
        f"{start_date.strftime('%d.%m.%Y')}"
        if start_date == end_date
        else f"{start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
    )

    lines = [
        f"📋 *Заказы на* | {period_text}",
        f"━━━━━━━━━━━━━━━━━━",
        f"📌 Всего: *{total_portions}* порций\n",
        f"• 🏢 Офис: *{office_portions}*",
        f"• 🏭 ПЦ 1: *{pc1_portions}*",
        f"• 📦 Склад: *{warehouse_portions}*"
    ]

    if unregistered_portions > 0:
        lines.append(f"• ❓ Незарегистрированные: *{unregistered_portions}*")

    return "\n".join(lines), total_portions


def generate_accounting_report_file(start_date, end_date, session):
    """
    Generate accounting Excel report (salary deductions).
    Returns (file_path, caption_text) or raises on error.
    """
    import locale
    try:
        locale.setlocale(locale.LC_NUMERIC, 'ru_RU.UTF-8')
    except Exception:
        pass

    reports_dir = ensure_reports_dir('accounting')
    now = datetime.now(CONFIG.timezone)

    if not start_date or not end_date:
        start_date = end_date = now.date()
    else:
        start_date = start_date if isinstance(start_date, date) else start_date.date()
        end_date = end_date if isinstance(end_date, date) else end_date.date()

    if start_date > end_date:
        start_date, end_date = end_date, start_date

    report_month = start_date.month
    report_year = start_date.year
    month_year = f"{MONTH_NAMES[report_month]} {report_year}"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Удержания за обеды"

    ws.append(["Список сотрудников на удержание обедов из ежемесячной премии"])
    ws.append([f"за {month_year} г."])
    ws.append([])
    ws.append(["", "удержание стоимости 1 обеда составляет", "150,00 руб. (без НДФЛ)"])
    ws.append(["", "", "172,41 руб. (с НДФЛ 13%)"])
    ws.append([])

    headers = [
        "Подразделение", "ФИО", "Кол-во обедов", "Должность",
        "Территория", "Дата приема", "Сумма удержания без НДФЛ", "Сумма удержания с НДФЛ"
    ]
    ws.append(headers)
    ws.auto_filter.ref = f"A{ws.max_row}:H{ws.max_row}"

    query = text('''
        SELECT
            COALESCE(u.department, 'Не указано') as department,
            u.full_name,
            SUM(o.quantity) as portions,
            COALESCE(u.position, 'Не указана') as position,
            COALESCE(u.city, 'Не указана') as city,
            u.employment_date as hire_date
        FROM orders o
        JOIN users u ON o.user_id = u.id
        WHERE o.target_date BETWEEN :start_date AND :end_date
          AND o.is_cancelled = FALSE
        GROUP BY u.id, u.department, u.full_name, u.position, u.city, u.employment_date
        ORDER BY u.department, u.full_name
    ''')

    rows = session.execute(query, {'start_date': start_date, 'end_date': end_date}).fetchall()

    total_portions = 0
    total_without_ndfl = 0
    total_with_ndfl = 0

    if not rows:
        ws.append(["Нет данных за выбранный период", "", "", "", "", "", "", ""])
    else:
        for row in rows:
            department, full_name, portions, position, city, hire_date = row

            hire_date_str = _format_hire_date(hire_date)

            amount_without_ndfl = portions * 150
            amount_with_ndfl = round(amount_without_ndfl / 0.87, 2)

            ws.append([
                department, full_name, portions, position, city,
                hire_date_str, amount_without_ndfl, amount_with_ndfl
            ])

            total_portions += portions
            total_without_ndfl += amount_without_ndfl
            total_with_ndfl += amount_with_ndfl

        ws.append(["ВСЕГО", "", total_portions, "", "", "", total_without_ndfl, total_with_ndfl])

    _apply_accounting_styles(ws)

    file_name = f"salary_deductions_{report_year}{report_month:02d}.xlsx"
    file_path = os.path.join(reports_dir, file_name)
    wb.save(file_path)

    def format_currency(amount):
        return f"{float(amount):,.2f}".replace(",", " ").replace(".", ",")

    caption = (
        f"📋 Отчет для удержаний из зарплаты\n"
        f"📅 Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n"
        f"🍽 Всего обедов: {total_portions}\n"
        f"💰 Сумма удержания: {format_currency(total_with_ndfl)} руб. (с НДФЛ)"
    )

    return file_path, file_name, caption


def generate_admin_report_file(start_date, end_date, session, is_daily=False):
    """
    Generate admin Excel report (orders by location).
    Returns (file_path, file_name, caption) or (None, None, no_data_message).
    """
    now = datetime.now(CONFIG.timezone)

    if not start_date or not end_date:
        month_start = now.replace(day=1).date()
        end_date = now.date()
        start_date = month_start
    else:
        if start_date > end_date:
            start_date, end_date = end_date, start_date

    reports_dir = ensure_reports_dir('admin')

    if is_daily:
        query = text("""
            SELECT o.target_date, u.full_name, COALESCE(u.location, 'Не указано') as location,
                   o.quantity, o.is_from_bitrix, o.created_at, o.bitrix_order_id
            FROM orders o JOIN users u ON o.user_id = u.id
            WHERE o.target_date = :target_date AND o.is_cancelled = FALSE
            ORDER BY o.target_date,
                     CASE WHEN o.bitrix_order_id IS NULL THEN CAST(o.created_at AS TEXT) ELSE o.bitrix_order_id END,
                     u.full_name
        """)
        all_orders = session.execute(query, {'target_date': start_date}).fetchall()
    else:
        query = text("""
            SELECT o.target_date, u.full_name, COALESCE(u.location, 'Не указано') as location,
                   o.quantity, o.is_from_bitrix, o.created_at, o.bitrix_order_id
            FROM orders o JOIN users u ON o.user_id = u.id
            WHERE o.target_date BETWEEN :start_date AND :end_date AND o.is_cancelled = FALSE
            ORDER BY o.target_date,
                     CASE WHEN o.bitrix_order_id IS NULL THEN CAST(o.created_at AS TEXT) ELSE o.bitrix_order_id END,
                     u.full_name
        """)
        all_orders = session.execute(query, {'start_date': start_date, 'end_date': end_date}).fetchall()

    if not all_orders:
        period_desc = (start_date.strftime("%d.%m.%Y") if is_daily
                       else f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
        return None, None, f"📊 На период {period_desc} заказов нет"

    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # "All orders" sheet
    ws_all = wb.create_sheet("Все заказы", 0)
    ws_all.append(["Дата обеда", "Номер заказа", "Сотрудник", "Локация", "Подпись", "Кол-во обедов", "Источник заказа"])
    ws_all.auto_filter.ref = "A1:G1"

    orders_by_date = {}
    for row in all_orders:
        key = row[0]
        orders_by_date.setdefault(key, []).append(row)

    for date_key in sorted(orders_by_date.keys()):
        for row in orders_by_date[date_key]:
            target_dt = row[0].strftime("%d.%m.%Y") if isinstance(row[0], date) else row[0]
            source = "Битрикс" if row[4] else "Бот"
            order_number = row[6] if row[6] is not None else ""
            ws_all.append([target_dt, order_number, row[1], row[2], "", row[3], source])

    # Per-location sheets
    for location in CONFIG.locations:
        if location == "ПЦ 2":
            continue
        ws = wb.create_sheet(location)
        ws.append(["Дата обеда", "Номер заказа", "Сотрудник", "Территориальный признак", "Подпись", "Кол-во обедов", "Источник заказа"])
        ws.auto_filter.ref = "A1:G1"

        if location == "Офис":
            loc_orders = [r for r in all_orders if r[2] in ["Офис", "ПЦ 2"]]
        else:
            loc_orders = [r for r in all_orders if r[2] == location]

        loc_by_date = {}
        for row in loc_orders:
            loc_by_date.setdefault(row[0], []).append(row)

        for date_key in sorted(loc_by_date.keys()):
            for row in loc_by_date[date_key]:
                target_dt = row[0].strftime("%d.%m.%Y") if isinstance(row[0], date) else row[0]
                source = "Битрикс" if row[4] else "Бот"
                order_number = row[6] if row[6] is not None else ""
                ws.append([target_dt, order_number, row[1], row[2], "", row[3], source])

    # Summary sheet
    ws_summary = wb.create_sheet("Итоги")
    ws_summary.append(["Локация", "Порции"])
    ws_summary.auto_filter.ref = "A1:B1"

    location_totals = {}
    for row in all_orders:
        loc = "Офис" if row[2] == "ПЦ 2" else row[2]
        location_totals[loc] = location_totals.get(loc, 0) + row[3]

    total = 0
    for loc, portions in sorted(location_totals.items(), key=lambda x: x[1], reverse=True):
        ws_summary.append([loc, portions])
        total += portions
    ws_summary.append(["ВСЕГО", total])

    # Styling
    bold_font = Font(bold=True)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = bold_font
        for col in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    timestamp = now.strftime("%Y%m%d_%H%M%S")
    file_name = f"admin_report_{timestamp}.xlsx"
    file_path = os.path.join(reports_dir, file_name)
    wb.save(file_path)

    if is_daily:
        caption = f"📅 Админ отчет за {start_date.strftime('%d.%m.%Y')}\n🍽 Всего порций: {total}"
    else:
        caption = f"📅 Админ отчет за период {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n🍽 Всего порций: {total}"

    return file_path, file_name, caption


# --- Helpers ---

def _format_hire_date(hire_date):
    """Format hire_date from various types to DD.MM.YYYY string."""
    if not hire_date:
        return "Не указана"
    if isinstance(hire_date, date):
        return hire_date.strftime("%d.%m.%Y")
    if isinstance(hire_date, datetime):
        return hire_date.date().strftime("%d.%m.%Y")
    if isinstance(hire_date, str):
        try:
            if '.' in hire_date:
                return datetime.strptime(hire_date, "%d.%m.%Y").strftime("%d.%m.%Y")
            elif '-' in hire_date:
                return datetime.strptime(hire_date, "%Y-%m-%d").strftime("%d.%m.%Y")
        except Exception:
            pass
    return "Не указана"


def _apply_accounting_styles(ws):
    """Apply formatting to accounting report worksheet."""
    bold_font = Font(bold=True)
    money_format = '# ##0.00'

    for row in ws.iter_rows(min_row=1, max_row=6):
        for cell in row:
            cell.font = bold_font

    for cell in ws[7]:
        cell.font = bold_font

    for cell in ws[ws.max_row]:
        cell.font = bold_font

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row):
        for cell in row[6:8]:
            if cell.value and isinstance(cell.value, (int, float)):
                cell.number_format = money_format
                cell.value = float(cell.value)

    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            length = len(str(cell.value or "")) * 1.2
            if cell.column_letter not in column_widths or length > column_widths[cell.column_letter]:
                column_widths[cell.column_letter] = length
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = min(width, 50)
